import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")

password = "hello"

passer = st.text_input("Enter password", type="password")

if passer == password:

    st.set_page_config(page_title="Complaints File Comparator", layout="centered")
    st.title("PT Files Comparator - Complaints")
    
    st.write("Upload the ECM File and the Power BI file.")
    
    
    selection = st.radio(
        "Choose your filter type:",
        ["PT", "Contracts", "Schools"],
        index = 0
    )
    
    
    # --- File Uploads ---
    st.markdown("<h2 style='text-align:center;color:#27AE60;'>Upload ECM File</h2>", unsafe_allow_html=True)
    st.write("Upload the ECM file as originally exported. In case you changed the filter type, reupload the ECM file.")
    uploaded_file1 = st.file_uploader("File 1", type=["xlsx", "xls", "xlsm", "xlsb"], key="file1")
    
    st.markdown("<h2 style='text-align:center;color:#2E86C1;'>Upload Power BI File</h2>", unsafe_allow_html=True)
    st.write("Upload the Power BI file as originally exported.")
    uploaded_file2 = st.file_uploader("File 2", type=["xlsx", "xls", "xlsm", "xlsb"], key="file2")
    
    if uploaded_file1 and uploaded_file2:
        try:
            # Read required columns
            file1 = pd.read_excel(uploaded_file1, header=2, sheet_name="Sheet1", usecols=["ComplaintNumber", "Fleet_Team", "Status", "Resolution_Range (INTERNAL-Aproval)"], dtype=str)
            file2 = pd.read_excel(uploaded_file2, usecols=["ComplaintNumber", "Status", "Closure Status"], dtype=str)
            file2 = file2[~file2["ComplaintNumber"].str.contains("applied", case=False, na=False)]
    
            file1_name = uploaded_file1.name
            file2_name = uploaded_file2.name
    
            if selection == "PT":
                file1 = file1[
                    (file1["Fleet_Team"].str.strip().str.lower().isin(["psv"]))
                ]
            elif selection == "Contracts":
                file1 = file1[
                    (file1["Fleet_Team"].str.strip().str.lower().isin(["contract"]))
                ]
            elif selection == "Schools":
                file1 = file1[
                    (file1["Fleet_Team"].str.strip().str.lower().isin(["sec"]))
                ]
    
            # Merge on ComplaintNumber
            merged = pd.merge(file1, file2, on="ComplaintNumber", how="outer", suffixes=('_file1', '_file2'))
    
            # Initialize counters
            MissingErrors = 0
            StateErrors = 0
            RangeErrors = 0
            error_IDs = set()
            analytics_data = []
    
            for _, row in merged.iterrows():
                num = row["ComplaintNumber"]
    
                if pd.isna(num):
                    continue
    
                # Missing in file1
                if pd.isna(row["Status_file1"]) and pd.isna(row["Resolution_Range (INTERNAL-Aproval)"]):
                    analytics_data.append({
                        "ComplaintNumber": num,
                        "Type": "Missing Complaint",
                        f"{file1_name}": "Missing",
                        f"{file2_name}": ""
                    })
                    MissingErrors += 1
                    error_IDs.add(num)
    
                # Missing in file2
                elif pd.isna(row["Status_file2"]) and pd.isna(row["Closure Status"]):
                    analytics_data.append({
                        "ComplaintNumber": num,
                        "Type": "Missing Complaint",
                        f"{file1_name}": "",
                        f"{file2_name}": "Missing"
                    })
                    MissingErrors += 1
                    error_IDs.add(num)
                else:
                    # Compare FileState
                    if row["Status_file1"] != row["Status_file2"]:
                        analytics_data.append({
                            "ComplaintNumber": num,
                            "Type": "State Mismatch",
                            f"{file1_name}": row["Status_file1"],
                            f"{file2_name}": row["Status_file2"]
                        })
                        StateErrors += 1
                        error_IDs.add(num)
    
                    # Compare Ranges
                    range1 = row["Resolution_Range (INTERNAL-Aproval)"]
                    range2 = row["Closure Status"]
                    if range1 != range2:
                        if not (range1 == "From 0 To 5 days" and range2 == "On Time"):
                            analytics_data.append({
                                "ComplaintNumber": num,
                                "Type": "Range Mismatch",
                                f"{file1_name}": range1,
                                f"{file2_name}": range2
                            })
                            RangeErrors += 1
                            error_IDs.add(num)
            valid_numbers = merged["ComplaintNumber"].dropna().unique()
            total_IDs = len(valid_numbers)
            total_error_IDs = len(error_IDs)
    
            if total_IDs > 0:
                accuracy = ((total_IDs - total_error_IDs) / total_IDs) * 100
                percent_error = (total_error_IDs / total_IDs) * 100
            else:
                accuracy = 0
                percent_error = 0     
    
            summary_table = [
                ["Total complaints for today", total_IDs, ""],
                ["Accuracy", f"{accuracy:.2f}%", ""],
                ["Files with mismatches", f"{percent_error:.2f}%", ""],
                ["Missing cases", MissingErrors, ""],
                ["State mismatches", StateErrors, ""],
                ["Range mismatches", RangeErrors, ""]
                ]
    
            df_comparison = pd.DataFrame(analytics_data)
    
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                sheet_name = "Comparison Result"
                df_comparison.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_table) + 3)
                worksheet = writer.sheets[sheet_name]
    
                # Header title
                worksheet.merge_cells('A1:D1')
                cell = worksheet['A1']
                cell.value = "Comparison Result"
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal="center")
                
                # Write summary table
                for i, row in enumerate(summary_table, start=3):
                    for j, value in enumerate(row, start=1):
                        cell = worksheet.cell(row=i, column=j, value=value)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
    
                # Adjust column widths
                for i, col in enumerate(worksheet.columns, start=1):
                    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
                    col_letter = get_column_letter(i)
                    worksheet.column_dimensions[col_letter].width = max_length + 4
    
                # Center-align analytics rows
                startrow = len(summary_table) + 3
                for row in worksheet.iter_rows(
                    min_row=startrow,
                    max_row=startrow + len(df_comparison) + 1,
                    min_col=1,
                    max_col=len(df_comparison.columns)
                ):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
    
            # --- Download Section ---
            st.markdown("<h4 style='color:#FF0000;'>Enter file name for download</h4>", unsafe_allow_html=True)
            filename = st.text_input("File name", value=f"{selection}_complaint_comparison_result.xlsx")
    
            st.download_button(
                label="💾 Download Comparison Excel",
                data=output.getvalue(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    
        except Exception as e:
    
            st.error(f"⚠️ Error: {str(e)}")
